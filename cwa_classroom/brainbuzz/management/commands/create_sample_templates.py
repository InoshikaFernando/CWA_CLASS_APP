"""Generate sample question upload template files (JSON, CSV, Excel).

Each file contains a DIFFERENT set of questions so teachers can upload all
three without hitting duplicate-question errors.
  JSON  → Foundational topics  (Addition, Subtraction, Multiplication, Division, Fractions, Arithmetic, BODMAS)
  CSV   → Intermediate topics  (Algebra, Geometry, Percentages, Ratios, Statistics, Number)
  Excel → Advanced / varied    (Area, Perimeter, Probability, Linear Equations, Integers, Prime Numbers, Square Roots)
"""

import json
import csv
import os

from django.core.management.base import BaseCommand

SAMPLES_DIR = os.path.join(os.path.dirname(__file__), 'samples')

# ── JSON question set ─────────────────────────────────────────────────────────
JSON_QUESTIONS = [
    # Addition
    {
        "topic": "Addition", "level": 1,
        "question_text": "What is 5 + 3?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "Count on from 5: 6, 7, 8. So 5 + 3 = 8.",
        "correct_short_answer": "",
        "answers": [
            {"text": "7", "is_correct": False, "order": 0},
            {"text": "8", "is_correct": True, "order": 1},
            {"text": "9", "is_correct": False, "order": 2},
            {"text": "6", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Addition", "level": 2,
        "question_text": "Fill in the blank: ___ + 45 = 100",
        "question_type": "fill_blank", "difficulty": 1, "points": 1,
        "explanation": "100 − 45 = 55. So the missing number is 55.",
        "correct_short_answer": "55",
        "answers": [],
    },
    {
        "topic": "Addition", "level": 3,
        "question_text": "What is 247 + 385?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "Add ones: 7+5=12, carry 1. Tens: 4+8+1=13, carry 1. Hundreds: 2+3+1=6. Answer: 632.",
        "correct_short_answer": "",
        "answers": [
            {"text": "622", "is_correct": False, "order": 0},
            {"text": "632", "is_correct": True, "order": 1},
            {"text": "642", "is_correct": False, "order": 2},
            {"text": "731", "is_correct": False, "order": 3},
        ],
    },
    # Subtraction
    {
        "topic": "Subtraction", "level": 2,
        "question_text": "What is 83 − 47?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "83 − 47 = 36.",
        "correct_short_answer": "",
        "answers": [
            {"text": "34", "is_correct": False, "order": 0},
            {"text": "36", "is_correct": True, "order": 1},
            {"text": "44", "is_correct": False, "order": 2},
            {"text": "46", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Subtraction", "level": 3,
        "question_text": "100 − 64 = 46.",
        "question_type": "true_false", "difficulty": 1, "points": 1,
        "explanation": "100 − 64 = 36, not 46. False.",
        "correct_short_answer": "",
        "answers": [
            {"text": "True", "is_correct": False, "order": 0},
            {"text": "False", "is_correct": True, "order": 1},
        ],
    },
    # Multiplication
    {
        "topic": "Multiplication", "level": 3,
        "question_text": "What is 9 × 6?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "9 × 6 = 54.",
        "correct_short_answer": "",
        "answers": [
            {"text": "48", "is_correct": False, "order": 0},
            {"text": "54", "is_correct": True, "order": 1},
            {"text": "56", "is_correct": False, "order": 2},
            {"text": "63", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Multiplication", "level": 4,
        "question_text": "What is 12 × 12?",
        "question_type": "short_answer", "difficulty": 1, "points": 1,
        "explanation": "12 × 12 = 144.",
        "correct_short_answer": "144",
        "answers": [],
    },
    # Division
    {
        "topic": "Division", "level": 3,
        "question_text": "What is 72 ÷ 8?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "72 ÷ 8 = 9. Check: 8 × 9 = 72.",
        "correct_short_answer": "",
        "answers": [
            {"text": "7", "is_correct": False, "order": 0},
            {"text": "8", "is_correct": False, "order": 1},
            {"text": "9", "is_correct": True, "order": 2},
            {"text": "10", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Division", "level": 3,
        "question_text": "48 ÷ 6 = 8.",
        "question_type": "true_false", "difficulty": 1, "points": 1,
        "explanation": "48 ÷ 6 = 8. Check: 6 × 8 = 48. True.",
        "correct_short_answer": "",
        "answers": [
            {"text": "True", "is_correct": True, "order": 0},
            {"text": "False", "is_correct": False, "order": 1},
        ],
    },
    # Fractions
    {
        "topic": "Fractions", "level": 5,
        "question_text": "What is 1/2 + 1/4?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Common denominator is 4. 1/2 = 2/4, so 2/4 + 1/4 = 3/4.",
        "correct_short_answer": "",
        "answers": [
            {"text": "3/4", "is_correct": True, "order": 0},
            {"text": "2/6", "is_correct": False, "order": 1},
            {"text": "1/8", "is_correct": False, "order": 2},
            {"text": "2/4", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Fractions", "level": 5,
        "question_text": "Is 3/6 equivalent to 1/2?",
        "question_type": "true_false", "difficulty": 1, "points": 1,
        "explanation": "3/6 simplifies to 1/2 by dividing both by 3. True.",
        "correct_short_answer": "",
        "answers": [
            {"text": "True", "is_correct": True, "order": 0},
            {"text": "False", "is_correct": False, "order": 1},
        ],
    },
    {
        "topic": "Fractions", "level": 6,
        "question_text": "Simplify the fraction 12/16 to its lowest terms.",
        "question_type": "short_answer", "difficulty": 2, "points": 2,
        "explanation": "GCF of 12 and 16 is 4. Dividing both by 4 gives 3/4.",
        "correct_short_answer": "3/4",
        "answers": [],
    },
    # Arithmetic
    {
        "topic": "Arithmetic", "level": 3,
        "question_text": "What is 7 × 8?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "7 × 8 = 56.",
        "correct_short_answer": "",
        "answers": [
            {"text": "54", "is_correct": False, "order": 0},
            {"text": "56", "is_correct": True, "order": 1},
            {"text": "58", "is_correct": False, "order": 2},
            {"text": "64", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Arithmetic", "level": 3,
        "question_text": "Division is the inverse (opposite) of multiplication.",
        "question_type": "true_false", "difficulty": 1, "points": 1,
        "explanation": "If a × b = c then c ÷ b = a. True.",
        "correct_short_answer": "",
        "answers": [
            {"text": "True", "is_correct": True, "order": 0},
            {"text": "False", "is_correct": False, "order": 1},
        ],
    },
    # BODMAS
    {
        "topic": "BODMAS", "level": 7,
        "question_text": "Evaluate: 3 + 4 × 2 − 1",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Multiplication first: 4 × 2 = 8. Then 3 + 8 − 1 = 10.",
        "correct_short_answer": "",
        "answers": [
            {"text": "10", "is_correct": True, "order": 0},
            {"text": "13", "is_correct": False, "order": 1},
            {"text": "14", "is_correct": False, "order": 2},
            {"text": "9", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "BODMAS", "level": 6,
        "question_text": "Evaluate: 12 ÷ (2 + 1) × 3",
        "question_type": "short_answer", "difficulty": 2, "points": 2,
        "explanation": "Brackets first: (2+1)=3. Then left to right: 12 ÷ 3 = 4, then 4 × 3 = 12.",
        "correct_short_answer": "12",
        "answers": [],
    },
]

# ── CSV question set ──────────────────────────────────────────────────────────
CSV_QUESTIONS = [
    # Algebra
    {
        "topic": "Algebra", "level": 7,
        "question_text": "Solve for x: 3x + 7 = 22",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Subtract 7: 3x = 15. Divide by 3: x = 5.",
        "correct_short_answer": "",
        "answers": [
            {"text": "3", "is_correct": False, "order": 0},
            {"text": "5", "is_correct": True, "order": 1},
            {"text": "7", "is_correct": False, "order": 2},
            {"text": "15", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Algebra", "level": 6,
        "question_text": "If 4x = 36, what is x?",
        "question_type": "short_answer", "difficulty": 1, "points": 1,
        "explanation": "Divide both sides by 4: x = 36 ÷ 4 = 9.",
        "correct_short_answer": "9",
        "answers": [],
    },
    {
        "topic": "Algebra", "level": 7,
        "question_text": "Solve for x: 2x + 5 = 13",
        "question_type": "short_answer", "difficulty": 2, "points": 2,
        "explanation": "Subtract 5: 2x = 8. Divide by 2: x = 4.",
        "correct_short_answer": "4",
        "answers": [],
    },
    {
        "topic": "Algebra", "level": 8,
        "question_text": "Which of the following is equivalent to 3(x + 4)?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Distribute: 3(x + 4) = 3x + 12.",
        "correct_short_answer": "",
        "answers": [
            {"text": "3x + 12", "is_correct": True, "order": 0},
            {"text": "3x + 4", "is_correct": False, "order": 1},
            {"text": "x + 12", "is_correct": False, "order": 2},
            {"text": "3x + 7", "is_correct": False, "order": 3},
        ],
    },
    # Geometry
    {
        "topic": "Geometry", "level": 4,
        "question_text": "What is the area of a rectangle with length 8 cm and width 5 cm?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "Area = length × width = 8 × 5 = 40 cm².",
        "correct_short_answer": "",
        "answers": [
            {"text": "13 cm²", "is_correct": False, "order": 0},
            {"text": "40 cm²", "is_correct": True, "order": 1},
            {"text": "26 cm²", "is_correct": False, "order": 2},
            {"text": "45 cm²", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Geometry", "level": 5,
        "question_text": "Fill in the blank: The angles in a triangle always add up to ___ degrees.",
        "question_type": "fill_blank", "difficulty": 2, "points": 1,
        "explanation": "Interior angles of any triangle sum to 180°.",
        "correct_short_answer": "180",
        "answers": [],
    },
    {
        "topic": "Geometry", "level": 5,
        "question_text": "What is the area of a triangle with base 10 cm and height 6 cm?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Area = ½ × base × height = ½ × 10 × 6 = 30 cm².",
        "correct_short_answer": "",
        "answers": [
            {"text": "60 cm²", "is_correct": False, "order": 0},
            {"text": "30 cm²", "is_correct": True, "order": 1},
            {"text": "16 cm²", "is_correct": False, "order": 2},
            {"text": "20 cm²", "is_correct": False, "order": 3},
        ],
    },
    # Percentages
    {
        "topic": "Percentages", "level": 5,
        "question_text": "What is 25% of 80?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "25% = 1/4. 80 ÷ 4 = 20.",
        "correct_short_answer": "",
        "answers": [
            {"text": "15", "is_correct": False, "order": 0},
            {"text": "20", "is_correct": True, "order": 1},
            {"text": "25", "is_correct": False, "order": 2},
            {"text": "40", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Percentages", "level": 6,
        "question_text": "A shirt costs $40. There is a 20% discount. What is the sale price?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "20% of $40 = $8 discount. Sale price = $40 − $8 = $32.",
        "correct_short_answer": "",
        "answers": [
            {"text": "$28", "is_correct": False, "order": 0},
            {"text": "$32", "is_correct": True, "order": 1},
            {"text": "$34", "is_correct": False, "order": 2},
            {"text": "$36", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Percentages", "level": 6,
        "question_text": "15 is what percentage of 60?",
        "question_type": "short_answer", "difficulty": 2, "points": 2,
        "explanation": "15 ÷ 60 × 100 = 25%.",
        "correct_short_answer": "25%|25",
        "answers": [],
    },
    # Ratios
    {
        "topic": "Ratios", "level": 6,
        "question_text": "Which of these is the simplified form of the ratio 15 : 25?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "GCF of 15 and 25 is 5. 15 ÷ 5 = 3, 25 ÷ 5 = 5. Answer: 3 : 5.",
        "correct_short_answer": "",
        "answers": [
            {"text": "5 : 3", "is_correct": False, "order": 0},
            {"text": "3 : 5", "is_correct": True, "order": 1},
            {"text": "1 : 5", "is_correct": False, "order": 2},
            {"text": "15 : 5", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Ratios", "level": 7,
        "question_text": "Write the ratio 12 : 18 in its simplest form.",
        "question_type": "short_answer", "difficulty": 2, "points": 2,
        "explanation": "GCF of 12 and 18 is 6. 12 ÷ 6 = 2, 18 ÷ 6 = 3. Answer: 2:3.",
        "correct_short_answer": "2:3|2 : 3",
        "answers": [],
    },
    # Statistics
    {
        "topic": "Statistics", "level": 7,
        "question_text": "What is the mean (average) of: 4, 6, 8, 10, 12?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Sum = 4+6+8+10+12 = 40. Count = 5. Mean = 40 ÷ 5 = 8.",
        "correct_short_answer": "",
        "answers": [
            {"text": "7", "is_correct": False, "order": 0},
            {"text": "8", "is_correct": True, "order": 1},
            {"text": "9", "is_correct": False, "order": 2},
            {"text": "10", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Statistics", "level": 7,
        "question_text": "What is the mode of: 3, 5, 5, 7, 8, 5, 2?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Mode = most frequent value. 5 appears 3 times. Mode = 5.",
        "correct_short_answer": "",
        "answers": [
            {"text": "3", "is_correct": False, "order": 0},
            {"text": "5", "is_correct": True, "order": 1},
            {"text": "7", "is_correct": False, "order": 2},
            {"text": "8", "is_correct": False, "order": 3},
        ],
    },
    # Number (decimal questions)
    {
        "topic": "Number", "level": 4,
        "question_text": "What is 0.5 + 0.25?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "0.50 + 0.25 = 0.75.",
        "correct_short_answer": "",
        "answers": [
            {"text": "0.70", "is_correct": False, "order": 0},
            {"text": "0.75", "is_correct": True, "order": 1},
            {"text": "0.55", "is_correct": False, "order": 2},
            {"text": "0.80", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Number", "level": 5,
        "question_text": "0.75 is equal to 3/4.",
        "question_type": "true_false", "difficulty": 1, "points": 1,
        "explanation": "3 ÷ 4 = 0.75. So 0.75 = 3/4. True.",
        "correct_short_answer": "",
        "answers": [
            {"text": "True", "is_correct": True, "order": 0},
            {"text": "False", "is_correct": False, "order": 1},
        ],
    },
]

# ── Excel question set ────────────────────────────────────────────────────────
EXCEL_QUESTIONS = [
    # Area
    {
        "topic": "Area", "level": 6,
        "question_text": "What is the area of a parallelogram with base 8 cm and height 5 cm?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Area of parallelogram = base × height = 8 × 5 = 40 cm².",
        "correct_short_answer": "",
        "answers": [
            {"text": "13 cm²", "is_correct": False, "order": 0},
            {"text": "40 cm²", "is_correct": True, "order": 1},
            {"text": "26 cm²", "is_correct": False, "order": 2},
            {"text": "20 cm²", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Area", "level": 7,
        "question_text": "What is the area of a circle with radius 7 cm? (Use π ≈ 3.14)",
        "question_type": "multiple_choice", "difficulty": 3, "points": 2,
        "explanation": "Area = π × r² = 3.14 × 7² = 3.14 × 49 ≈ 153.86 cm².",
        "correct_short_answer": "",
        "answers": [
            {"text": "43.96 cm²", "is_correct": False, "order": 0},
            {"text": "153.86 cm²", "is_correct": True, "order": 1},
            {"text": "98 cm²", "is_correct": False, "order": 2},
            {"text": "21.98 cm²", "is_correct": False, "order": 3},
        ],
    },
    # Perimeter
    {
        "topic": "Perimeter", "level": 3,
        "question_text": "Fill in the blank: Perimeter of a square with side 6 cm = ___ cm.",
        "question_type": "fill_blank", "difficulty": 1, "points": 1,
        "explanation": "Perimeter of square = 4 × side = 4 × 6 = 24 cm.",
        "correct_short_answer": "24",
        "answers": [],
    },
    {
        "topic": "Perimeter", "level": 4,
        "question_text": "What is the perimeter of a triangle with sides 5 cm, 7 cm, and 9 cm?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "Perimeter = sum of all sides = 5 + 7 + 9 = 21 cm.",
        "correct_short_answer": "",
        "answers": [
            {"text": "19 cm", "is_correct": False, "order": 0},
            {"text": "21 cm", "is_correct": True, "order": 1},
            {"text": "23 cm", "is_correct": False, "order": 2},
            {"text": "63 cm", "is_correct": False, "order": 3},
        ],
    },
    # Probability
    {
        "topic": "Probability", "level": 6,
        "question_text": "Probability of an event always lies between 0 and 1 inclusive.",
        "question_type": "true_false", "difficulty": 1, "points": 1,
        "explanation": "P = 0 means impossible, P = 1 means certain. All probabilities lie in [0, 1].",
        "correct_short_answer": "",
        "answers": [
            {"text": "True", "is_correct": True, "order": 0},
            {"text": "False", "is_correct": False, "order": 1},
        ],
    },
    {
        "topic": "Probability", "level": 7,
        "question_text": "What is the probability of rolling a 4 on a fair six-sided die?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "There is 1 favourable outcome out of 6 equally likely outcomes. P = 1/6.",
        "correct_short_answer": "",
        "answers": [
            {"text": "1/4", "is_correct": False, "order": 0},
            {"text": "1/6", "is_correct": True, "order": 1},
            {"text": "1/3", "is_correct": False, "order": 2},
            {"text": "4/6", "is_correct": False, "order": 3},
        ],
    },
    # Linear Equations
    {
        "topic": "Linear Equations", "level": 8,
        "question_text": "Solve: 2x − 3 = x + 5",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Subtract x from both sides: x − 3 = 5. Add 3: x = 8.",
        "correct_short_answer": "",
        "answers": [
            {"text": "2", "is_correct": False, "order": 0},
            {"text": "8", "is_correct": True, "order": 1},
            {"text": "4", "is_correct": False, "order": 2},
            {"text": "−8", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Linear Equations", "level": 8,
        "question_text": "Solve: 3(x − 2) = 9",
        "question_type": "short_answer", "difficulty": 2, "points": 2,
        "explanation": "Expand: 3x − 6 = 9. Add 6: 3x = 15. Divide by 3: x = 5.",
        "correct_short_answer": "5",
        "answers": [],
    },
    # Integers
    {
        "topic": "Integers", "level": 5,
        "question_text": "What is −5 + 8?",
        "question_type": "multiple_choice", "difficulty": 1, "points": 1,
        "explanation": "Start at −5 and count up 8: −5 + 8 = 3.",
        "correct_short_answer": "",
        "answers": [
            {"text": "−3", "is_correct": False, "order": 0},
            {"text": "3", "is_correct": True, "order": 1},
            {"text": "13", "is_correct": False, "order": 2},
            {"text": "−13", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Integers", "level": 6,
        "question_text": "What is −3 × −4?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Negative × Negative = Positive. −3 × −4 = 12.",
        "correct_short_answer": "",
        "answers": [
            {"text": "−12", "is_correct": False, "order": 0},
            {"text": "12", "is_correct": True, "order": 1},
            {"text": "7", "is_correct": False, "order": 2},
            {"text": "−7", "is_correct": False, "order": 3},
        ],
    },
    # Prime Numbers
    {
        "topic": "Prime Numbers", "level": 5,
        "question_text": "Which of these is a prime number?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "23 is divisible only by 1 and itself. 21=3×7, 25=5×5, 27=3×9.",
        "correct_short_answer": "",
        "answers": [
            {"text": "21", "is_correct": False, "order": 0},
            {"text": "23", "is_correct": True, "order": 1},
            {"text": "25", "is_correct": False, "order": 2},
            {"text": "27", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Prime Numbers", "level": 5,
        "question_text": "1 is a prime number.",
        "question_type": "true_false", "difficulty": 2, "points": 1,
        "explanation": "By definition, a prime number has exactly 2 distinct factors. 1 has only 1 factor. False.",
        "correct_short_answer": "",
        "answers": [
            {"text": "True", "is_correct": False, "order": 0},
            {"text": "False", "is_correct": True, "order": 1},
        ],
    },
    # Square Roots
    {
        "topic": "Square Roots", "level": 7,
        "question_text": "What is √144?",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "12 × 12 = 144. So √144 = 12.",
        "correct_short_answer": "",
        "answers": [
            {"text": "11", "is_correct": False, "order": 0},
            {"text": "12", "is_correct": True, "order": 1},
            {"text": "13", "is_correct": False, "order": 2},
            {"text": "14", "is_correct": False, "order": 3},
        ],
    },
    {
        "topic": "Square Roots", "level": 7,
        "question_text": "What is √225?",
        "question_type": "short_answer", "difficulty": 2, "points": 2,
        "explanation": "15 × 15 = 225. So √225 = 15.",
        "correct_short_answer": "15",
        "answers": [],
    },
    # BODMAS (extra)
    {
        "topic": "BODMAS", "level": 8,
        "question_text": "Evaluate: (2 + 3)² − 4 × 3 + 1",
        "question_type": "multiple_choice", "difficulty": 3, "points": 2,
        "explanation": "Brackets: (2+3)=5. Power: 5²=25. Multiply: 4×3=12. Then: 25 − 12 + 1 = 14.",
        "correct_short_answer": "",
        "answers": [
            {"text": "14", "is_correct": True, "order": 0},
            {"text": "10", "is_correct": False, "order": 1},
            {"text": "38", "is_correct": False, "order": 2},
            {"text": "26", "is_correct": False, "order": 3},
        ],
    },
    # Fractions (extra)
    {
        "topic": "Fractions", "level": 6,
        "question_text": "Convert 1 3/4 to an improper fraction.",
        "question_type": "short_answer", "difficulty": 2, "points": 2,
        "explanation": "(1 × 4) + 3 = 7. Denominator stays 4. Answer: 7/4.",
        "correct_short_answer": "7/4",
        "answers": [],
    },
    # Arithmetic
    {
        "topic": "Arithmetic", "level": 5,
        "question_text": "Round 4,672 to the nearest hundred.",
        "question_type": "multiple_choice", "difficulty": 2, "points": 1,
        "explanation": "Tens digit is 7 (≥5), so round up. 4,672 → 4,700.",
        "correct_short_answer": "",
        "answers": [
            {"text": "4,600", "is_correct": False, "order": 0},
            {"text": "4,700", "is_correct": True, "order": 1},
            {"text": "4,000", "is_correct": False, "order": 2},
            {"text": "5,000", "is_correct": False, "order": 3},
        ],
    },
]


def _build_json_payload(questions):
    result = []
    for q in questions:
        entry = {
            "topic": q["topic"],
            "level": q["level"],
            "question_text": q["question_text"],
            "question_type": q["question_type"],
            "difficulty": q["difficulty"],
            "points": q["points"],
            "explanation": q["explanation"],
        }
        if q["correct_short_answer"]:
            entry["correct_short_answer"] = q["correct_short_answer"]
        if q["answers"]:
            entry["answers"] = q["answers"]
        result.append(entry)
    return {"subject": "maths", "questions": result}


def _build_csv_rows(questions):
    max_answers = max((len(q["answers"]) for q in questions), default=0)
    answer_cols = []
    for i in range(1, max_answers + 1):
        answer_cols += [f"answer{i}", f"is_correct{i}"]

    headers = [
        "topic", "level", "question_text", "question_type",
        "difficulty", "points", "explanation", "correct_short_answer",
    ] + answer_cols

    rows = []
    for q in questions:
        row = {
            "topic": q["topic"],
            "level": q["level"],
            "question_text": q["question_text"],
            "question_type": q["question_type"],
            "difficulty": q["difficulty"],
            "points": q["points"],
            "explanation": q["explanation"],
            "correct_short_answer": q["correct_short_answer"],
        }
        for i, ans in enumerate(q["answers"], start=1):
            row[f"answer{i}"] = ans["text"]
            row[f"is_correct{i}"] = "true" if ans["is_correct"] else "false"
        rows.append(row)

    return headers, rows


class Command(BaseCommand):
    help = "Generate sample question upload templates (JSON, CSV, Excel) in management/commands/samples/"

    def handle(self, *args, **options):
        os.makedirs(SAMPLES_DIR, exist_ok=True)

        self._write_json()
        self._write_csv()
        self._write_excel()

        self.stdout.write(self.style.SUCCESS(f"Sample templates written to {SAMPLES_DIR}"))

    def _write_json(self):
        path = os.path.join(SAMPLES_DIR, "sample_maths_questions.json")
        payload = _build_json_payload(JSON_QUESTIONS)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        self.stdout.write(f"  JSON  ({len(JSON_QUESTIONS)} questions): {path}")

    def _write_csv(self):
        path = os.path.join(SAMPLES_DIR, "sample_maths_questions.csv")
        headers, rows = _build_csv_rows(CSV_QUESTIONS)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
        self.stdout.write(f"  CSV   ({len(CSV_QUESTIONS)} questions): {path}")

    def _write_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.stderr.write("openpyxl not installed — skipping Excel template")
            return

        path = os.path.join(SAMPLES_DIR, "sample_maths_questions.xlsx")
        headers, rows = _build_csv_rows(EXCEL_QUESTIONS)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
        header_align = Alignment(horizontal="center", wrap_text=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        alt_fill = PatternFill(fill_type="solid", fgColor="EAF0FB")
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
                cell.alignment = Alignment(wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        ws.freeze_panes = "A2"

        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = 20

        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["G"].width = 40

        wb.save(path)
        self.stdout.write(f"  Excel ({len(EXCEL_QUESTIONS)} questions): {path}")
