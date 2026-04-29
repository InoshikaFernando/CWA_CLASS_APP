"""Question file parsers for JSON, CSV, and Excel formats.

Supports importing questions in multiple formats with automatic validation.
Each parser returns a list of question dictionaries ready for database insertion.
"""

import json
import csv
import io
from abc import ABC, abstractmethod
from typing import List, Tuple, Dict, Any
from decimal import Decimal


class BaseQuestionParser(ABC):
    """Abstract base class for question file parsers."""

    # Valid question types
    VALID_QUESTION_TYPES = {
        'multiple_choice', 'true_false', 'short_answer', 'fill_blank',
        'write_code', 'calculation'
    }

    # Valid difficulty levels
    VALID_DIFFICULTIES = {1, 2, 3}

    def __init__(self):
        self.errors = []

    @abstractmethod
    def parse(self, file_obj) -> List[Dict]:
        """Parse file and return list of question dictionaries.

        Returns:
            List of question dicts with keys: question_text, question_type,
            difficulty, level_id (or topic_level_id for coding), topic_id,
            answers, etc.
        """
        pass

    def validate_question(self, q_dict: Dict) -> Tuple[bool, List[str]]:
        """Validate a single question dictionary.

        Returns:
            (is_valid, errors_list)
        """
        errors = []

        # Required fields
        if not q_dict.get('question_text', '').strip():
            errors.append('question_text is required and must be non-empty')

        if q_dict.get('question_type') not in self.VALID_QUESTION_TYPES:
            errors.append(
                f"question_type must be one of {self.VALID_QUESTION_TYPES}, "
                f"got {q_dict.get('question_type')}"
            )

        if q_dict.get('difficulty') not in self.VALID_DIFFICULTIES:
            errors.append(
                f"difficulty must be 1, 2, or 3, got {q_dict.get('difficulty')}"
            )

        # Check for topic_id or topic_level_id or topic_name (name resolved later)
        if (not q_dict.get('topic_id') and not q_dict.get('topic_level_id')
                and not q_dict.get('topic_name')):
            errors.append('Either topic_id (maths) or topic_level_id (coding) is required')

        # Check for level_id if using explicit topic_id (topic_name + level_number is also valid)
        if q_dict.get('topic_id') and not q_dict.get('level_id'):
            errors.append('level_id is required when using topic_id')

        # Validate answers if MCQ/TF
        if q_dict.get('question_type') in ('multiple_choice', 'true_false'):
            answers = q_dict.get('answers', [])
            if not answers:
                errors.append('Answers required for multiple_choice and true_false')

            correct_count = sum(1 for a in answers if a.get('is_correct'))
            if correct_count == 0:
                errors.append('At least one answer must be marked as correct')

            if q_dict.get('question_type') == 'true_false' and len(answers) != 2:
                errors.append('true_false must have exactly 2 answers')

        # Validate short answer / fill blank
        if q_dict.get('question_type') in ('short_answer', 'fill_blank'):
            if not q_dict.get('correct_short_answer', '').strip():
                errors.append(
                    f"correct_short_answer is required for {q_dict.get('question_type')}"
                )

        return len(errors) == 0, errors


class JSONQuestionParser(BaseQuestionParser):
    """Parser for JSON format question files.

    Expected format:
    {
        "subject": "maths" or "coding",
        "questions": [
            {
                "question_text": "...",
                "question_type": "multiple_choice",
                "difficulty": 2,
                "topic": "Fractions",  # topic name (will lookup ID)
                "level": 5,  # year level 1-8 (will lookup ID)
                "answers": [
                    {"text": "...", "is_correct": true, "order": 1},
                    ...
                ],
                ...
            }
        ]
    }
    """

    def parse(self, file_obj) -> List[Dict]:
        """Parse JSON file and return questions."""
        try:
            # Read and decode file
            if isinstance(file_obj, io.TextIOBase):
                content = file_obj.read()
            else:
                content = file_obj.read().decode('utf-8')

            data = json.loads(content)
        except json.JSONDecodeError as e:
            self.errors.append(f"Invalid JSON: {str(e)}")
            return []
        except Exception as e:
            self.errors.append(f"Error reading file: {str(e)}")
            return []

        if not isinstance(data, dict):
            self.errors.append("JSON root must be an object")
            return []

        questions = data.get('questions', [])
        if not questions:
            self.errors.append("No 'questions' key found in JSON")
            return []

        parsed_questions = []
        for idx, q in enumerate(questions):
            parsed_q = self._parse_question(q, idx)
            if parsed_q:
                parsed_questions.append(parsed_q)

        return parsed_questions

    def _parse_question(self, q: Dict, idx: int) -> Dict | None:
        """Parse a single question from JSON."""
        try:
            parsed = {
                'question_text': q.get('question_text', '').strip(),
                'question_type': q.get('question_type', '').lower(),
                'difficulty': int(q.get('difficulty', 1)),
                'points': int(q.get('points', 1)),
                'explanation': q.get('explanation', '').strip(),
                'correct_short_answer': q.get('correct_short_answer', '').strip() or None,
            }

            # Parse topic and level
            topic_name = q.get('topic', '').strip()
            level_num = q.get('level')

            if not topic_name or level_num is None:
                self.errors.append(
                    f"Question {idx}: 'topic' and 'level' are required"
                )
                return None

            parsed['topic_name'] = topic_name
            parsed['level_number'] = int(level_num)

            # Parse answers
            answers = q.get('answers', [])
            if answers:
                parsed_answers = []
                for ans_idx, ans in enumerate(answers):
                    parsed_ans = {
                        'text': ans.get('text', '').strip() or ans.get('answer_text', '').strip(),
                        'is_correct': bool(ans.get('is_correct', False)),
                        'order': int(ans.get('order', ans_idx)),
                    }
                    if not parsed_ans['text']:
                        self.errors.append(
                            f"Question {idx}, Answer {ans_idx}: text is required"
                        )
                        return None
                    parsed_answers.append(parsed_ans)
                parsed['answers'] = parsed_answers
            else:
                parsed['answers'] = []

            # Validate
            is_valid, errors = self.validate_question(parsed)
            if not is_valid:
                for err in errors:
                    self.errors.append(f"Question {idx}: {err}")
                return None

            return parsed

        except Exception as e:
            self.errors.append(f"Question {idx}: Error parsing - {str(e)}")
            return None


class CSVQuestionParser(BaseQuestionParser):
    """Parser for CSV format question files.

    Expected headers:
    topic, level, question_text, question_type, difficulty, [points], [explanation],
    answer1, is_correct1, [answer2], [is_correct2], ...

    Example:
    topic,level,question_text,question_type,difficulty,answer1,is_correct1,answer2,is_correct2,answer3,is_correct3
    Fractions,5,"What is 1/2 + 1/4?",multiple_choice,2,3/4,true,2/4,false,1/4,false
    """

    def parse(self, file_obj) -> List[Dict]:
        """Parse CSV file and return questions."""
        try:
            if isinstance(file_obj, io.TextIOBase):
                content = file_obj.read()
            else:
                content = file_obj.read().decode('utf-8')

            reader = csv.DictReader(io.StringIO(content))
            if not reader.fieldnames:
                self.errors.append("CSV is empty or has no headers")
                return []

            parsed_questions = []
            for row_idx, row in enumerate(reader, start=2):  # Start at 2 (header is row 1)
                parsed_q = self._parse_row(row, row_idx)
                if parsed_q:
                    parsed_questions.append(parsed_q)

            return parsed_questions

        except Exception as e:
            self.errors.append(f"Error reading CSV: {str(e)}")
            return []

    def _parse_row(self, row: Dict[str, str], row_idx: int) -> Dict | None:
        """Parse a single CSV row into a question."""
        try:
            # Required fields
            topic = (row.get('topic') or '').strip()
            level = (row.get('level') or '').strip()
            question_text = (row.get('question_text') or '').strip()
            question_type = (row.get('question_type') or '').strip().lower()
            difficulty = (row.get('difficulty') or '1').strip()

            if not all([topic, level, question_text, question_type]):
                self.errors.append(
                    f"Row {row_idx}: Missing required fields (topic, level, question_text, question_type)"
                )
                return None

            parsed = {
                'topic_name': topic,
                'level_number': int(level),
                'question_text': question_text,
                'question_type': question_type,
                'difficulty': int(difficulty),
                'points': int((row.get('points') or '1').strip()),
                'explanation': (row.get('explanation') or '').strip(),
                'correct_short_answer': (row.get('correct_short_answer') or '').strip() or None,
            }

            # Parse answer pairs (answer1, is_correct1, answer2, is_correct2, ...)
            answers = []
            order = 0
            for key, value in sorted(row.items()):
                if key.startswith('answer'):
                    # Extract number: answer1 -> 1
                    try:
                        ans_num = int(key.replace('answer', ''))
                    except ValueError:
                        continue

                    answer_text = (value or '').strip()
                    if not answer_text:
                        continue

                    # Get corresponding is_correct value
                    is_correct_key = f'is_correct{ans_num}'
                    is_correct_val = (row.get(is_correct_key) or '').strip().lower()
                    is_correct = is_correct_val in ('true', '1', 'yes', 'y')

                    answers.append({
                        'text': answer_text,
                        'is_correct': is_correct,
                        'order': order,
                    })
                    order += 1

            if answers:
                parsed['answers'] = answers
            else:
                parsed['answers'] = []

            # Validate
            is_valid, errors = self.validate_question(parsed)
            if not is_valid:
                for err in errors:
                    self.errors.append(f"Row {row_idx}: {err}")
                return None

            return parsed

        except ValueError as e:
            self.errors.append(f"Row {row_idx}: Invalid numeric value - {str(e)}")
            return None
        except Exception as e:
            self.errors.append(f"Row {row_idx}: Error parsing - {str(e)}")
            return None


class ExcelQuestionParser(BaseQuestionParser):
    """Parser for Excel format question files.

    Uses openpyxl to read Excel spreadsheets.
    Expected sheet name: "Questions" (first sheet used if not found)
    Expected headers: Same as CSV format
    """

    def parse(self, file_obj) -> List[Dict]:
        """Parse Excel file and return questions."""
        try:
            import openpyxl
        except ImportError:
            self.errors.append(
                "openpyxl is required for Excel support. Install: pip install openpyxl"
            )
            return []

        try:
            workbook = openpyxl.load_workbook(file_obj)

            # Try to find "Questions" sheet, otherwise use first sheet
            if 'Questions' in workbook.sheetnames:
                sheet = workbook['Questions']
            else:
                sheet = workbook.active

            if not sheet or sheet.max_row < 2:
                self.errors.append("Sheet is empty or has no data rows")
                return []

            # Convert sheet to CSV-like format for parsing
            parsed_questions = []
            headers = None

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    # Header row
                    headers = [str(h).strip() if h else '' for h in row]
                else:
                    # Data row
                    row_dict = {
                        headers[i]: str(val).strip() if val is not None else ''
                        for i, val in enumerate(row)
                        if i < len(headers)
                    }
                    parsed_q = self._parse_row(row_dict, row_idx)
                    if parsed_q:
                        parsed_questions.append(parsed_q)

            workbook.close()
            return parsed_questions

        except Exception as e:
            self.errors.append(f"Error reading Excel file: {str(e)}")
            return []

    def _parse_row(self, row: Dict[str, str], row_idx: int) -> Dict | None:
        """Parse a single Excel row (reuse CSV parser logic)."""
        # Reuse CSV parser's row parsing logic
        csv_parser = CSVQuestionParser()
        result = csv_parser._parse_row(row, row_idx)
        if csv_parser.errors:
            self.errors.extend(csv_parser.errors)
        return result


def get_parser(file_format: str) -> BaseQuestionParser | None:
    """Factory function to get the appropriate parser for a file format.

    Args:
        file_format: 'json', 'csv', or 'excel'

    Returns:
        Parser instance or None if format not supported
    """
    parsers = {
        'json': JSONQuestionParser,
        'csv': CSVQuestionParser,
        'excel': ExcelQuestionParser,
    }
    
    parser_class = parsers.get(file_format.lower())
    return parser_class() if parser_class else None
